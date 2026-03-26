"""
Fetches all archived URLs for leg.state.co.us from the Wayback Machine CDX API,
checks each one for 200/404, then analyzes patterns in the live URLs.

Usage: python check_urls.py
"""

import sys
import time
import re
from collections import Counter
from urllib.parse import urlparse

try:
    import requests
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "requests"])
    import requests

CDX_URL = (
    "http://web.archive.org/cdx/search/cdx"
    "?url=leg.state.co.us/*"
    "&output=text"
    "&fl=original"
    "&collapse=urlkey"
    "&limit=50000"
)

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; url-checker/1.0)"}
DELAY = 0.3  # seconds between requests — be polite


def fetch_cdx(retries=3):
    print("Fetching URL list from Wayback Machine CDX API...")
    print("(This can take a minute -- streaming large response)\n")
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(CDX_URL, headers=HEADERS, timeout=120, stream=True)
            resp.raise_for_status()
            lines = []
            for line in resp.iter_lines():
                if line:
                    lines.append(line.decode("utf-8").strip())
            print(f"  {len(lines)} URLs retrieved.\n")
            return lines
        except requests.exceptions.Timeout:
            print(f"  Timeout on attempt {attempt}/{retries}, retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"  Error on attempt {attempt}/{retries}: {e}")
            time.sleep(5)
    print("Failed to fetch CDX data after all retries.")
    sys.exit(1)


def check_urls(urls):
    live, dead, errors = [], [], []
    total = len(urls)

    for i, url in enumerate(urls, 1):
        try:
            r = requests.head(url, headers=HEADERS, timeout=10, allow_redirects=True)
            status = r.status_code
        except requests.exceptions.SSLError:
            # Retry with SSL verification off for old gov sites
            try:
                r = requests.head(url, headers=HEADERS, timeout=10, allow_redirects=True, verify=False)
                status = r.status_code
            except Exception as e:
                errors.append((url, str(e)))
                print(f"  [{i:>5}/{total}] ERROR   {url}")
                continue
        except Exception as e:
            errors.append((url, str(e)))
            print(f"  [{i:>5}/{total}] ERROR   {url}")
            continue

        if status == 200:
            live.append(url)
            print(f"  [{i:>5}/{total}] {status} OK    {url}")
        else:
            dead.append((url, status))
            print(f"  [{i:>5}/{total}] {status}       {url}")

        time.sleep(DELAY)

    return live, dead, errors


def relative_path(url):
    parsed = urlparse(url)
    path = parsed.path.rstrip("/") or "/"
    return path


def analyze_patterns(live_urls):
    paths = [relative_path(u) for u in live_urls]

    # Extract first path segment (top-level "directories")
    top_level = Counter()
    for p in paths:
        parts = p.strip("/").split("/")
        if parts and parts[0]:
            top_level[parts[0]] += 1

    # Extract second path segment
    second_level = Counter()
    for p in paths:
        parts = p.strip("/").split("/")
        if len(parts) >= 2 and parts[1]:
            second_level[f"/{parts[0]}/{parts[1]}"] += 1

    # File extensions
    ext_counter = Counter()
    for p in paths:
        m = re.search(r"\.(\w+)$", p)
        ext_counter[m.group(1) if m else "(no ext)"] += 1

    return paths, top_level, second_level, ext_counter


def save_results(live, dead, errors, paths, top_level, second_level, ext_counter):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Sheet 1: All URLs ---
    ws_all = wb.active
    ws_all.title = "All URLs"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill_green = PatternFill("solid", start_color="375623")
    header_fill_red   = PatternFill("solid", start_color="7B1818")
    header_fill_gray  = PatternFill("solid", start_color="595959")

    def write_header(ws, cols, fill):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    write_header(ws_all, ["Status", "URL", "Relative Path", "Top-Level Segment"], header_fill_green)

    green_fill = PatternFill("solid", start_color="E2EFDA")
    red_fill   = PatternFill("solid", start_color="FFDCDC")
    gray_fill  = PatternFill("solid", start_color="F2F2F2")

    for url in sorted(live):
        p = relative_path(url)
        top = p.strip("/").split("/")[0] if p.strip("/") else "/"
        row = ws_all.max_row + 1
        ws_all.append([200, url, p, f"/{top}"])
        for cell in ws_all[row]:
            cell.fill = green_fill
            cell.font = Font(name="Arial")

    for url, status in sorted(dead):
        p = relative_path(url)
        top = p.strip("/").split("/")[0] if p.strip("/") else "/"
        row = ws_all.max_row + 1
        ws_all.append([status, url, p, f"/{top}"])
        for cell in ws_all[row]:
            cell.fill = red_fill
            cell.font = Font(name="Arial")

    for url, err in errors:
        p = relative_path(url)
        top = p.strip("/").split("/")[0] if p.strip("/") else "/"
        row = ws_all.max_row + 1
        ws_all.append(["ERROR", url, p, f"/{top}"])
        for cell in ws_all[row]:
            cell.fill = gray_fill
            cell.font = Font(name="Arial")

    ws_all.column_dimensions["A"].width = 10
    ws_all.column_dimensions["B"].width = 80
    ws_all.column_dimensions["C"].width = 60
    ws_all.column_dimensions["D"].width = 30
    ws_all.auto_filter.ref = f"A1:D{ws_all.max_row}"
    ws_all.freeze_panes = "A2"

    # --- Sheet 2: Patterns ---
    ws_pat = wb.create_sheet("Patterns")
    write_header(ws_pat, ["Segment / Pattern", "Count", "Type"], header_fill_gray)

    for seg, count in top_level.most_common(50):
        ws_pat.append([f"/{seg}", count, "Top-level"])
        ws_pat[ws_pat.max_row][0].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][1].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][2].font = Font(name="Arial")

    ws_pat.append([])
    for seg, count in second_level.most_common(50):
        ws_pat.append([seg, count, "Second-level"])
        ws_pat[ws_pat.max_row][0].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][1].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][2].font = Font(name="Arial")

    ws_pat.append([])
    for ext, count in ext_counter.most_common():
        ws_pat.append([f".{ext}", count, "Extension"])
        ws_pat[ws_pat.max_row][0].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][1].font = Font(name="Arial")
        ws_pat[ws_pat.max_row][2].font = Font(name="Arial")

    ws_pat.column_dimensions["A"].width = 40
    ws_pat.column_dimensions["B"].width = 12
    ws_pat.column_dimensions["C"].width = 20

    # --- Sheet 3: Live only ---
    ws_live = wb.create_sheet("Live")
    write_header(ws_live, ["URL", "Relative Path", "Top-Level Segment"], header_fill_green)
    for url in sorted(live):
        p = relative_path(url)
        top = p.strip("/").split("/")[0] if p.strip("/") else "/"
        ws_live.append([url, p, f"/{top}"])
        for cell in ws_live[ws_live.max_row]:
            cell.font = Font(name="Arial")
    ws_live.column_dimensions["A"].width = 80
    ws_live.column_dimensions["B"].width = 60
    ws_live.column_dimensions["C"].width = 30
    ws_live.auto_filter.ref = f"A1:C{ws_live.max_row}"
    ws_live.freeze_panes = "A2"

    out = "url_results.xlsx"
    wb.save(out)
    print(f"\nSaved results to {out}")


def main():
    # Optionally load URLs from a file if you already have the CDX results
    if len(sys.argv) > 1:
        fname = sys.argv[1]
        print(f"Loading URLs from {fname}...")
        with open(fname) as f:
            urls = [u.strip() for u in f if u.strip()]
        print(f"  {len(urls)} URLs loaded.\n")
    else:
        urls = fetch_cdx()

    # Optionally limit for testing: urls = urls[:100]

    print(f"Checking {len(urls)} URLs (this may take a while)...\n")
    live, dead, errors = check_urls(urls)

    print(f"\n--- Results ---")
    print(f"  Live:   {len(live)}")
    print(f"  Dead:   {len(dead)}")
    print(f"  Errors: {len(errors)}")

    print("\nAnalyzing patterns in live URLs...")
    paths, top_level, second_level, ext_counter = analyze_patterns(live)

    print("\n=== TOP-LEVEL PATH SEGMENTS ===")
    for seg, count in top_level.most_common(20):
        print(f"  {count:>5}  /{seg}")

    print("\n=== FILE EXTENSIONS ===")
    for ext, count in ext_counter.most_common(10):
        print(f"  {count:>5}  .{ext}")

    save_results(live, dead, errors, paths, top_level, second_level, ext_counter)


if __name__ == "__main__":
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    main()