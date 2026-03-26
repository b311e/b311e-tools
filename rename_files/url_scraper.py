import re
import sys
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def default_filename(soup):
    title = soup.title.string.strip() if soup.title and soup.title.string else "untitled"
    slug = re.sub(r"[^\w\s-]", "", title)
    slug = re.sub(r"[\s_]+", "-", slug).strip("-").lower()
    slug = slug[:60]
    timestamp = datetime.now().strftime("%Y-%m-%d-%H%M")
    return f"{timestamp}_{slug}.xlsx"


def scrape_urls(page_url, output_file=None):
    response = requests.get(page_url, headers={"User-Agent": "Mozilla/5.0"})
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    base_url = f"{urlparse(page_url).scheme}://{urlparse(page_url).netloc}"

    if output_file is None:
        output_file = default_filename(soup)

    wb = Workbook()
    ws = wb.active
    ws.title = "URLs"

    # Header row
    for col, label in enumerate(["Link Text", "URL"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, name="Arial")

    # Data rows
    row = 2
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith(("#", "mailto:", "javascript:")):
            continue
        full_url = urljoin(base_url, href)
        link_text = a.get_text(strip=True) or "[no text]"
        ws.cell(row=row, column=1, value=link_text).font = Font(name="Arial")
        url_cell = ws.cell(row=row, column=2, value=full_url)
        url_cell.font = Font(name="Arial", color="0563C1", underline="single")
        url_cell.hyperlink = full_url
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 70

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_file)
    print(f"Saved {row - 2} URLs to {output_file}")


if __name__ == "__main__":
    url = input("Enter the URL to scrape: ").strip()
    if not url:
        print("No URL provided. Exiting.")
        sys.exit(1)
    out = input("Output filename [press Enter for auto-generated name]: ").strip() or None
    scrape_urls(url, out)